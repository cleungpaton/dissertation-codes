"""
h3_regression_corrected.py
--------------------------
Applies manual review corrections to automated scores, then runs the
H3 logistic regression on corrected accuracy data.

Logic:
  1. Load teacher and student feature CSVs (automated scores)
  2. Load manual review Excel sheets
  3. For any (problem_id, run_id) in the review sheet, override
     the 'correct' column with YOUR_VERDICT (Y=1, N=0)
  4. Run the same logistic regression as before

Usage:
    python3 h3_regression_corrected.py \
        --teacher teacher_features.jsonl \
        --student student_features.jsonl \
        --teacher-review teacher_manual_review.xlsx \
        --student-review student_manual_review_FULL.xlsx
"""

import argparse
import numpy as np
import pandas as pd
import openpyxl
from statsmodels.formula.api import logit


FEATURES = ['V1', 'V2', 'V3', 'V4', 'V5']


def load_features(path):
    """Load features CSV file."""
    df = pd.read_csv(path)
    required = ['problem_id', 'model_name', 'run_id', 'correct', 'complexity_level']
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"ERROR: Missing columns in {path}: {missing}")
        print(f"Available: {list(df.columns)}")
        exit(1)
    return df


def load_manual_review(path):
    """Load manual review Excel and return dict of (problem_id, run_id) -> corrected_score."""
    wb = openpyxl.load_workbook(path)
    ws = wb['Sheet1']

    corrections = {}
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 1).value      # problem_id
        rid = ws.cell(r, 2).value      # run_id
        verdict = ws.cell(r, 7).value  # YOUR_VERDICT: Y or N

        if pid is None or verdict is None:
            continue

        corrected = 1 if verdict.strip().upper() == 'Y' else 0
        corrections[(pid, int(rid))] = corrected

    return corrections


def apply_corrections(df, corrections, model_label):
    """Override automated 'correct' with manual verdicts where available."""
    changes = 0
    reviewed = 0

    for idx, row in df.iterrows():
        key = (row['problem_id'], int(row['run_id']))
        if key in corrections:
            reviewed += 1
            old = int(row['correct'])
            new = corrections[key]
            if old != new:
                df.at[idx, 'correct'] = new
                changes += 1

    print(f"  {model_label}: {reviewed} rows reviewed, {changes} scores changed")
    return df


def main():
    parser = argparse.ArgumentParser(description='H3 regression with corrected accuracy')
    parser.add_argument('--teacher', required=True, help='Path to teacher features CSV')
    parser.add_argument('--student', required=True, help='Path to student features CSV')
    parser.add_argument('--teacher-review', required=True, help='Path to teacher manual review Excel')
    parser.add_argument('--student-review', required=True, help='Path to student manual review Excel')
    args = parser.parse_args()

    # Load feature data
    print("Loading feature data...")
    teacher = load_features(args.teacher)
    print(f"  Teacher: {len(teacher)} rows")
    student = load_features(args.student)
    print(f"  Student: {len(student)} rows")

    # Load manual reviews
    print("\nLoading manual reviews...")
    teacher_corrections = load_manual_review(args.teacher_review)
    print(f"  Teacher review: {len(teacher_corrections)} entries")
    student_corrections = load_manual_review(args.student_review)
    print(f"  Student review: {len(student_corrections)} entries")

    # Apply corrections
    print("\nApplying corrections...")
    teacher = apply_corrections(teacher, teacher_corrections, "Teacher")
    student = apply_corrections(student, student_corrections, "Student")

    # Show corrected accuracy by level for sanity check
    print("\n--- Corrected Accuracy by Level ---")
    for label, df in [("Teacher", teacher), ("Student", student)]:
        print(f"\n  {label}:")
        for level in sorted(df['complexity_level'].unique()):
            subset = df[df['complexity_level'] == level]
            acc = subset['correct'].mean()
            n = len(subset)
            print(f"    Level {level}: {acc:.3f} ({int(subset['correct'].sum())}/{n})")
        overall = df['correct'].mean()
        print(f"    Overall: {overall:.3f} ({int(df['correct'].sum())}/{len(df)})")

    # Combine and run regression
    combined = pd.concat([teacher, student], ignore_index=True)
    combined['is_student'] = (combined['model_name'] == 'student').astype(int)
    combined['correct'] = combined['correct'].astype(int)
    combined['complexity'] = combined['complexity_level'].astype(int)
    combined['student_x_complexity'] = combined['is_student'] * combined['complexity']

    # Accuracy gap table
    print("\n--- Accuracy Gap (Teacher - Student) by Level ---")
    for level in sorted(combined['complexity'].unique()):
        t_acc = combined[(combined['complexity'] == level) & (combined['is_student'] == 0)]['correct'].mean()
        s_acc = combined[(combined['complexity'] == level) & (combined['is_student'] == 1)]['correct'].mean()
        gap = t_acc - s_acc
        print(f"  Level {level}: Teacher={t_acc:.3f}, Student={s_acc:.3f}, Gap={gap:+.3f}")

    # Fit logistic regression
    print("\n" + "=" * 65)
    print("LOGISTIC REGRESSION (CORRECTED): correct ~ is_student * complexity")
    print("=" * 65)

    model = logit('correct ~ is_student + complexity + student_x_complexity', data=combined).fit(disp=0)
    print(model.summary())

    # Key results
    coefs = model.params
    pvals = model.pvalues

    print("\n" + "=" * 65)
    print("KEY RESULTS (CORRECTED)")
    print("=" * 65)

    print(f"\nIntercept (b0):              {coefs['Intercept']:+.4f}  (p={pvals['Intercept']:.4f})")
    print(f"is_student (b1):             {coefs['is_student']:+.4f}  (p={pvals['is_student']:.4f})")
    print(f"complexity (b2):             {coefs['complexity']:+.4f}  (p={pvals['complexity']:.4f})")
    print(f"is_student x complexity (b3):{coefs['student_x_complexity']:+.4f}  (p={pvals['student_x_complexity']:.4f})")

    b3 = coefs['student_x_complexity']
    p3 = pvals['student_x_complexity']

    print(f"\n--- Interpretation ---")
    if p3 < 0.05:
        if b3 < 0:
            print(f"b3 is NEGATIVE and SIGNIFICANT (p={p3:.4f}).")
            print("-> The student's accuracy drops faster than the teacher's")
            print("   as complexity increases. H3 IS SUPPORTED.")
        else:
            print(f"b3 is POSITIVE and SIGNIFICANT (p={p3:.4f}).")
            print("-> Gap narrows with complexity. H3 NOT supported.")
    else:
        print(f"b3 is NOT SIGNIFICANT (p={p3:.4f}).")
        print("-> No evidence the accuracy gap widens with complexity.")
        print("   H3 is not supported at the 0.05 level.")

    # Odds ratios
    print(f"\n--- Odds Ratios ---")
    print(f"is_student:              OR = {np.exp(coefs['is_student']):.4f}")
    print(f"complexity:              OR = {np.exp(coefs['complexity']):.4f}")
    print(f"is_student x complexity: OR = {np.exp(coefs['student_x_complexity']):.4f}")

    # Model fit
    print(f"\n--- Model Fit ---")
    print(f"Log-likelihood:    {model.llf:.2f}")
    print(f"AIC:               {model.aic:.2f}")
    print(f"BIC:               {model.bic:.2f}")
    print(f"Pseudo R-squared:  {model.prsquared:.4f}")
    print(f"N observations:    {model.nobs:.0f}")

    # Save
    conf = model.conf_int()
    results = pd.DataFrame({
        'Coefficient': coefs.index,
        'Estimate': coefs.values.round(4),
        'Std_Error': model.bse.values.round(4),
        'z_value': model.tvalues.values.round(4),
        'p_value': pvals.values.round(4),
        'CI_lower': conf[0].values.round(4),
        'CI_upper': conf[1].values.round(4),
        'Odds_Ratio': np.exp(coefs.values).round(4),
    })
    results.to_csv('h3_regression_corrected.csv', index=False)
    print(f"\nResults saved to: h3_regression_corrected.csv")


if __name__ == '__main__':
    main()
